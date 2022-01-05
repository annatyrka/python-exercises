import openpyxl
import re

wb = openpyxl.load_workbook(
    '/Users/annatyrka/PythonPractice/excel/Impero - BRM entities and submitters list.xlsx')
sheet = wb['For Impero AT']
sheet2 = wb['Split']
email_regex = re.compile(r'''(
    [a-zA-Z0-9.%+-]+
    @
    [a-zA-Z0-9.-]+
    (\.[a-zA-Z]{2,5})
)''', re.VERBOSE)
matches = []
for row_num in range(2, sheet.max_row):
    responsible_cell = sheet.cell(row=row_num, column=4).value
    # print(responsible_cell)
    print(type(responsible_cell))
    for groups in email_regex.findall(responsible_cell):
        matches.append(groups[0])
n = 2
for person in matches:
    print(person)
    sheet2['D' + str(n)].value = person
    n += 1
wb.save('excel/separted_emails.xlsx')

  var moviesNolan = watchList
  .map(movie => {
    return {
      director: movie.Director,
      rating: movie.imdbRating
  };
  })
  .filter(movie => {
    return movie.Director === "Christopher Nolan"
  });
