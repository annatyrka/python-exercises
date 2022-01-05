import openpyxl
import csv
import os

os.chdir('/Users/annatyrka/PythonPractice/Excel')

for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):
        continue
    print(f"Opening file{excel_file}...")
    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        print(sheet_name)

        sheet = wb[sheet_name]
        print(f"    Opening sheet {sheet.title}...")

        csv_file = open(str(excel_file[:-5]) +
                        "_" + str(sheet.title)+".csv", "w")
        csv_writer = csv.writer(csv_file)

        for rowNum in range(1, sheet.max_row + 1):
            rowData = []
            for colNum in range(1, sheet.max_column + 1):
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                # print(rowData)
            csv_writer.writerow(rowData)
        csv_file.close()
excel_file.close()
