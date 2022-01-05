import openpyxl

file1 = open('excel/mad_libs.txt')
file2 = open('excel/mad_libs2.txt')

wb = openpyxl.Workbook()
sheet = wb.active

i = 1
for line in file1.readlines():
    sheet.cell(row=i, column=1).value = line
    i += 1

i = 1
for line in file2.readlines():
    sheet.cell(row=i, column=2).value = line
    i += 1

wb.save('excel/txt_to_excel.xlsx')
