import openpyxl

wb = openpyxl.load_workbook('excel/txt_to_excel.xlsx')
sheet = wb.active

for col_num in range(1, sheet.max_column + 1):
    file1 = open('excel/file'+str(col_num)+'.txt', 'w')
    for row_num in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=col_num).value is None:
            file1.write("")
        else:
            file1.write(sheet.cell(row=row_num, column=col_num).value)
    file1.close()
