import openpyxl

wb = openpyxl.load_workbook('excel/produceSales.xlsx')
sheet = wb['Sheet']

price_updates = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in price_updates:
        sheet.cell(row=row_num, column=2).value = price_updates[produce_name]

wb.save('excel/updated_produceSales.xlsx')
