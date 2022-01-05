import openpyxl
import re

wb = openpyxl.load_workbook('/Users/annatyrka/PythonPractice/Book1.xlsx')
sheet = wb['Sheet1']
sheet2 = wb['Sheet2']
email_regex = re.compile(r'''(
    [a-zA-Z0-9.%+-]+
    @
    [a-zA-Z0-9.-]+
    (\.[a-zA-Z]{2,5})
)''', re.VERBOSE)
matches = []
for row_num in range(2, sheet.max_row+1):
    responsible_cell = sheet.cell(row=row_num, column=6).value
    print(responsible_cell)
    print(type(responsible_cell))
    for groups in email_regex.findall(responsible_cell):
        matches.append(groups[0])
print(matches)
n = 2
for person in matches:
    print(person)
    sheet2['A' + str(n)].value = person
    n += 1
wb.save('excel/separted_names.xlsx')
