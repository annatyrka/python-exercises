import openpyxl
import sys

n = int(sys.argv[1])
m = int(sys.argv[2])

wb = openpyxl.load_workbook('excel/'+sys.argv[3])
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for i in range(1, n):
    for j in range(1, sheet.max_column+1):
        cell_value = sheet.cell(row=i, column=j).value
        new_sheet.cell(row=i, column=j).value = cell_value

for i in range(1, m+1):
    new_sheet.cell(row=new_sheet.max_row + n + i-1, column=1).value = ""

for i in range(n, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        cell_value = sheet.cell(row=i, column=j).value
        new_sheet.cell(row=i+m, column=j).value = cell_value

new_wb.save('excel/multiplication_v2.xlsx')
