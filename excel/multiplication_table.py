import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys

n = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
for row_num in range(1, n+1):
    sheet.cell(row=row_num + 1, column=1).value = row_num
for col_num in range(1, n+1):
    sheet.cell(row=1, column=col_num + 1).value = col_num

fontObj = Font(bold=True)
for i in range(2, n+2):
    sheet['A'+str(i)].font = fontObj
    sheet[get_column_letter(i)+str(1)].font = fontObj

for row_num in range(2, sheet.max_row+1):
    for col_num in range(2, sheet.max_column+1):
        sheet.cell(row=row_num, column=col_num).value = sheet.cell(
            row=row_num, column=1).value * sheet.cell(row=1, column=col_num).value

wb.save('multiplication.xlsx')
# :'A'+str(n+1)
# get_column_letter(n+1)+str(1)
