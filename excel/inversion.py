import openpyxl

# list1 = []

wb = openpyxl.load_workbook('excel/inversion.xlsx')
sheet = wb.active

data_sheet = []

for row_num in range(1, sheet.max_row+1):
    row_list = []
    for col_num in range(1, sheet.max_column+1):
        row_list.append(sheet.cell(row=row_num, column=col_num).value)
    data_sheet.append(row_list)

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for y in range(len(data_sheet)):
    for x in range(len(data_sheet[y])):
        new_sheet.cell(row=x+1, column=y+1).value = data_sheet[y][x]

new_wb.save('excel/inversion2.xlsx')
