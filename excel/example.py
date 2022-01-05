import openpyxl
import os

os.chdir('excel')

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)  # The workbook's sheets' names.

sheet2 = wb['Sheet2']
print(sheet2.title)

another_sheet = wb.active  # Get the active sheet
print(another_sheet)

sheet1 = wb['Sheet1']
cell = sheet1['A1']
print(cell.value)

cell2 = sheet1['B1']
print('Row %s, column %s is %s' % (cell.row, cell.column, cell.value))
print('Cell %s is %s' % (cell.coordinate, cell.value))

print(cell2.coordinate)

b2 = sheet1.cell(row=2, column=2)
print(b2.column)

for i in range(1, 8):
    print(i, sheet1.cell(row=i, column=2).value)

print(sheet1.max_column)
