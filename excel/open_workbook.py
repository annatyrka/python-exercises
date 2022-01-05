import openpyxl
import os

wb = openpyxl.Workbook()  # Create a blank workbook
print(wb.sheetnames)

sheet = wb.active
sheet.title
sheet.title = "Sheet1"  # Change title
print(wb.sheetnames)
wb.save('created_file.xlsx')

wb['Sheet1']['A1'] = "Hello, wordld!"
print(wb['Sheet1']['A1'].value)
